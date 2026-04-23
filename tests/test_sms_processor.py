from __future__ import annotations

import re
from pathlib import Path
from unittest.mock import MagicMock, patch, call

import pytest
from cryptography.fernet import Fernet


# ---------------------------------------------------------------------------
# Helpers used across tests
# ---------------------------------------------------------------------------

def make_engine_mock(fetchall_return=None, rowcount=0):
    """Return a mock SQLAlchemy engine whose context managers work."""
    engine = MagicMock()
    conn = MagicMock()
    conn.execute.return_value.fetchall.return_value = fetchall_return or []
    conn.execute.return_value.rowcount = rowcount

    # engine.connect() as conn
    engine.connect.return_value.__enter__ = MagicMock(return_value=conn)
    engine.connect.return_value.__exit__ = MagicMock(return_value=False)

    # engine.begin() as conn
    engine.begin.return_value.__enter__ = MagicMock(return_value=conn)
    engine.begin.return_value.__exit__ = MagicMock(return_value=False)

    return engine, conn


def make_config(dry_run=False):
    cfg = MagicMock()
    cfg.get.return_value = {
        'blasta_ini': 'secrets/BLASTA.ini',
        'blasta_key': 'secrets/BLASTA.key',
        'max_retries': 3,
        'dry_run': dry_run,
    }
    return cfg


# ---------------------------------------------------------------------------
# load_templates_from_excel
# ---------------------------------------------------------------------------

def test_load_templates_from_excel_reads_all_languages(tmp_path):
    """Each xlsx file produces templates for both weeks."""
    import openpyxl
    from scripts.seed_sms_templates import load_templates_from_excel

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Arm', 'Wk 8 SMS', 'Wk 11 SMS'])
    ws.append(['HIV Risk Assessment', 'Week 8 message', 'Week 11 message'])
    wb.save(tmp_path / 'English.xlsx')

    templates = load_templates_from_excel(str(tmp_path))

    assert len(templates) == 2
    assert templates[0] == {
        'arm': 'HIV Risk Assessment',
        'language': 'English',
        'week': 8,
        'message_text': 'Week 8 message',
        'has_placeholder': False,
    }
    assert templates[1]['week'] == 11


def test_load_templates_detects_placeholder(tmp_path):
    import openpyxl
    from scripts.seed_sms_templates import load_templates_from_excel

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Arm', 'Wk 8 SMS', 'Wk 11 SMS'])
    ws.append(['Default appointment setting', 'Your appt is [date here]', 'Follow up on [date]'])
    wb.save(tmp_path / 'English.xlsx')

    templates = load_templates_from_excel(str(tmp_path))
    assert all(t['has_placeholder'] for t in templates)


def test_load_templates_skips_blank_arms(tmp_path):
    import openpyxl
    from scripts.seed_sms_templates import load_templates_from_excel

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Arm', 'Wk 8 SMS', 'Wk 11 SMS'])
    ws.append([None, 'msg', 'msg'])  # blank arm — should be skipped
    ws.append(['HIV Risk Assessment', 'msg', 'msg'])
    wb.save(tmp_path / 'English.xlsx')

    templates = load_templates_from_excel(str(tmp_path))
    assert len(templates) == 2


# ---------------------------------------------------------------------------
# _load_blasta_creds
# ---------------------------------------------------------------------------

def test_load_blasta_creds_roundtrip(tmp_path):
    from modules.sms_processor import _load_blasta_creds

    key = Fernet.generate_key()
    cipher = Fernet(key)
    encrypted = cipher.encrypt(b'hunter2').decode()

    ini = tmp_path / 'BLASTA.ini'
    ini.write_text(f"Username=myuser\nPassword={encrypted}\n")
    key_file = tmp_path / 'BLASTA.key'
    key_file.write_text(key.decode())

    username, password = _load_blasta_creds(str(ini), str(key_file))
    assert username == 'myuser'
    assert password == 'hunter2'


def test_load_blasta_creds_missing_username_raises(tmp_path):
    from modules.sms_processor import _load_blasta_creds

    key = Fernet.generate_key()
    cipher = Fernet(key)
    encrypted = cipher.encrypt(b'pw').decode()

    ini = tmp_path / 'BLASTA.ini'
    ini.write_text(f"Password={encrypted}\n")  # no Username
    key_file = tmp_path / 'BLASTA.key'
    key_file.write_text(key.decode())

    with pytest.raises(KeyError, match='Username'):
        _load_blasta_creds(str(ini), str(key_file))


def test_load_blasta_creds_missing_password_raises(tmp_path):
    from modules.sms_processor import _load_blasta_creds

    key = Fernet.generate_key()
    ini = tmp_path / 'BLASTA.ini'
    ini.write_text("Username=myuser\n")  # no Password
    key_file = tmp_path / 'BLASTA.key'
    key_file.write_text(key.decode())

    with pytest.raises(KeyError, match='Password'):
        _load_blasta_creds(str(ini), str(key_file))


# ---------------------------------------------------------------------------
# _substitute_placeholder
# ---------------------------------------------------------------------------

def test_substitute_placeholder_replaces_bracket_text():
    from modules.sms_processor import _substitute_placeholder
    result = _substitute_placeholder('Your appt is [insert date]', '25/12/2025')
    assert result == 'Your appt is 25/12/2025'


def test_substitute_placeholder_no_bracket_unchanged():
    from modules.sms_processor import _substitute_placeholder
    msg = 'No placeholders here'
    assert _substitute_placeholder(msg, '25/12/2025') == msg


def test_substitute_placeholder_none_date_unchanged():
    from modules.sms_processor import _substitute_placeholder
    msg = 'Your appt is [insert date]'
    assert _substitute_placeholder(msg, None) == msg


def test_substitute_placeholder_invalid_date_unchanged(caplog):
    from modules.sms_processor import _substitute_placeholder
    import logging
    with caplog.at_level(logging.WARNING):
        result = _substitute_placeholder('Appt: [date]', 'not-a-date')
    assert result == 'Appt: [date]'
    assert 'Invalid appointment date' in caplog.text


# ---------------------------------------------------------------------------
# BlastaClient
# ---------------------------------------------------------------------------

def test_blasta_client_sends_successfully():
    from modules.sms_processor import BlastaClient

    with patch('modules.sms_processor.requests.post') as mock_post:
        # First call: get_token; second call: send_sms
        mock_post.side_effect = [
            MagicMock(status_code=200, json=lambda: {'access_token': 'tok123'}),
            MagicMock(status_code=200, json=lambda: {'msg_id': 'M001', 'status_code': '201'}),
        ]
        client = BlastaClient('user', 'pass', max_retries=3)
        result = client.send('0700000001', 'Hello')

    assert result['msg_id'] == 'M001'
    assert mock_post.call_count == 2


def test_blasta_client_send_logs_warning_when_no_msg_id(caplog):
    """When API returns empty msg_id, send still succeeds but logs a warning."""
    import logging
    from modules.sms_processor import BlastaClient

    with patch('modules.sms_processor.requests.post') as mock_post:
        mock_post.side_effect = [
            MagicMock(status_code=200, json=lambda: {'access_token': 'tok123'}),
            MagicMock(status_code=200, raise_for_status=MagicMock(),
                      json=lambda: {'status_code': '201', 'msg_id': ''}),
        ]
        client = BlastaClient('user', 'pass', max_retries=3)
        with caplog.at_level(logging.WARNING):
            result = client.send('0700000001', 'Hello')

    assert result.get('msg_id') in (None, '')
    assert 'msg_id' in caplog.text.lower() or 'provider' in caplog.text.lower()


def test_blasta_client_refreshes_token_on_401():
    from modules.sms_processor import BlastaClient

    token_resp = MagicMock(status_code=200, json=lambda: {'access_token': 'newtok'})
    token_resp.raise_for_status = MagicMock()
    fail_401 = MagicMock(status_code=401, json=lambda: {})
    fail_401.raise_for_status = MagicMock()
    success_resp = MagicMock(status_code=200, json=lambda: {'msg_id': 'M002'})
    success_resp.raise_for_status = MagicMock()

    with patch('modules.sms_processor.requests.post') as mock_post:
        # get_token → 401 send → get_token refresh → success send
        mock_post.side_effect = [token_resp, fail_401, token_resp, success_resp]

        client = BlastaClient('user', 'pass', max_retries=3)
        result = client.send('0700000001', 'Hello')

    assert result['msg_id'] == 'M002'


def test_blasta_client_raises_after_max_retries():
    from modules.sms_processor import BlastaClient

    import requests as req_lib

    token_resp = MagicMock(status_code=200, json=lambda: {'access_token': 'tok'})
    token_resp.raise_for_status = MagicMock()

    with patch('modules.sms_processor.requests.post') as mock_post:
        with patch('modules.sms_processor.time.sleep'):
            mock_post.side_effect = [
                token_resp,
                req_lib.RequestException('timeout'),
                req_lib.RequestException('timeout'),
                req_lib.RequestException('timeout'),
            ]

            client = BlastaClient('user', 'pass', max_retries=3)
            with pytest.raises(req_lib.RequestException):
                client.send('0700000001', 'Hello')


# ---------------------------------------------------------------------------
# SmsProcessor.sync_queue
# ---------------------------------------------------------------------------

def test_sync_queue_executes_three_statements():
    """sync_queue runs 2 INSERT statements + 1 UPDATE for opt-outs."""
    from modules.sms_processor import SmsProcessor

    engine, conn = make_engine_mock(rowcount=2)
    processor = SmsProcessor(config=make_config(), engine=engine)
    processor.sync_queue()

    assert conn.execute.call_count == 3


def test_sync_queue_uses_countrycode_from_config():
    """Countrycode is read from config, not hardcoded."""
    from modules.sms_processor import SmsProcessor

    engine, conn = make_engine_mock(rowcount=2)

    cfg = MagicMock()
    cfg.get.return_value = {
        'blasta_ini': 'secrets/BLASTA.ini',
        'blasta_key': 'secrets/BLASTA.key',
        'max_retries': 3,
        'dry_run': False,
        'countrycode': '2',  # non-Uganda country
    }

    processor = SmsProcessor(config=cfg, engine=engine)
    processor.sync_queue()

    executed_sql = [str(c.args[0]) for c in conn.execute.call_args_list]
    # The hardcoded '1' must NOT appear in any SQL — must be a bind parameter
    assert not any("countrycode = '1'" in s for s in executed_sql), \
        "countrycode '1' is hardcoded — must be passed as a bind parameter"


# ---------------------------------------------------------------------------
# SmsProcessor.get_due_messages
# ---------------------------------------------------------------------------

def test_get_due_messages_returns_list_of_dicts():
    from modules.sms_processor import SmsProcessor

    row = MagicMock()
    row._asdict.return_value = {
        'id': 1, 'subjid': 'IBIS001', 'mobile_number': '0700000001',
        'arm_text': 'HIV Risk Assessment', 'language': 'English',
        'week': 8, 'appointment_date': None,
    }
    engine, conn = make_engine_mock(fetchall_return=[row])
    processor = SmsProcessor(config=make_config(), engine=engine)
    result = processor.get_due_messages()

    assert len(result) == 1
    assert result[0]['subjid'] == 'IBIS001'


# ---------------------------------------------------------------------------
# SmsProcessor.send_due_messages — dry run
# ---------------------------------------------------------------------------

def test_send_due_messages_dry_run_skips_all():
    from modules.sms_processor import SmsProcessor

    row = MagicMock()
    row._asdict.return_value = {
        'id': 1, 'subjid': 'IBIS001', 'mobile_number': '0700000001',
        'arm_text': 'HIV Risk Assessment', 'language': 'English',
        'week': 8, 'appointment_date': None,
    }
    template_row = MagicMock()
    template_row.message_text = 'Please visit the clinic.'
    template_row.has_placeholder = False

    engine, conn = make_engine_mock()
    # get_due_messages returns one row; _resolve_template returns template
    conn.execute.return_value.fetchall.return_value = [row]
    conn.execute.return_value.fetchone.return_value = template_row

    processor = SmsProcessor(config=make_config(dry_run=True), engine=engine)
    result = processor.send_due_messages()

    assert result.sent == 0
    assert result.skipped == 1
    assert result.failed == 0


# ---------------------------------------------------------------------------
# SmsProcessor.send_due_messages — missing template skips row
# ---------------------------------------------------------------------------

def test_send_due_messages_missing_template_skips():
    from modules.sms_processor import SmsProcessor

    row = MagicMock()
    row._asdict.return_value = {
        'id': 2, 'subjid': 'IBIS002', 'mobile_number': '0700000002',
        'arm_text': 'Unknown Arm', 'language': 'Klingon',
        'week': 8, 'appointment_date': None,
    }
    engine, conn = make_engine_mock()
    conn.execute.return_value.fetchall.return_value = [row]
    conn.execute.return_value.fetchone.return_value = None  # no template found

    processor = SmsProcessor(config=make_config(), engine=engine)
    result = processor.send_due_messages()

    assert result.skipped == 1
    assert result.sent == 0


# ---------------------------------------------------------------------------
# SendSms stage
# ---------------------------------------------------------------------------

def test_send_sms_stage_returns_success_result():
    from stages.send_sms import SendSms
    from modules.sms_processor import SendResult

    engine, _ = make_engine_mock()
    config = make_config()

    stage = SendSms(config=config, engine=engine)

    with patch('stages.send_sms.SmsProcessor') as MockProcessor:
        MockProcessor.return_value.run.return_value = SendResult(sent=5, failed=0, skipped=1)
        result = stage.run()

    assert result.success is True
    assert result.rows_written == 5
    assert result.errors == []
    assert result.metadata['sent'] == 5
    assert result.metadata['skipped'] == 1


def test_send_sms_stage_returns_failure_on_failed_sends():
    from stages.send_sms import SendSms
    from modules.sms_processor import SendResult

    engine, _ = make_engine_mock()
    stage = SendSms(config=make_config(), engine=engine)

    with patch('stages.send_sms.SmsProcessor') as MockProcessor:
        MockProcessor.return_value.run.return_value = SendResult(
            sent=3, failed=2, skipped=0,
            failures=[
                {'subjid': 'X1', 'mobile_number': '07001', 'week': 8, 'error': 'timeout'},
                {'subjid': 'X2', 'mobile_number': '07002', 'week': 11, 'error': 'bad number'},
            ],
        )
        result = stage.run()

    assert result.success is False
    assert result.rows_written == 3
    assert len(result.errors) == 2
    assert 'subjid=X1' in result.errors[0]


# ---------------------------------------------------------------------------
# BlastaClient.check_dlr
# ---------------------------------------------------------------------------

def test_check_dlr_returns_status_on_success():
    from modules.sms_processor import BlastaClient

    with patch('modules.sms_processor.requests.post') as mock_post:
        token_resp = MagicMock(status_code=200, json=lambda: {'access_token': 'tok'})
        token_resp.raise_for_status = MagicMock()
        dlr_resp = MagicMock(status_code=200, json=lambda: {
            'msg_id': '12345', 'status': 'DELIVERED',
            'submitted_at': '2026-04-23 09:00:00', 'status_code': 200,
        })
        dlr_resp.raise_for_status = MagicMock()
        mock_post.side_effect = [token_resp, dlr_resp]

        client = BlastaClient('user', 'pass')
        status = client.check_dlr('12345')

    assert status == 'DELIVERED'


def test_check_dlr_returns_not_found_on_404():
    from modules.sms_processor import BlastaClient

    with patch('modules.sms_processor.requests.post') as mock_post:
        token_resp = MagicMock(status_code=200, json=lambda: {'access_token': 'tok'})
        token_resp.raise_for_status = MagicMock()
        not_found = MagicMock(status_code=404)
        not_found.raise_for_status = MagicMock()
        mock_post.side_effect = [token_resp, not_found]

        client = BlastaClient('user', 'pass')
        status = client.check_dlr('99999')

    assert status == 'NOT_FOUND'


def test_check_dlr_refreshes_token_on_401():
    from modules.sms_processor import BlastaClient

    with patch('modules.sms_processor.requests.post') as mock_post:
        token_resp = MagicMock(status_code=200, json=lambda: {'access_token': 'tok'})
        token_resp.raise_for_status = MagicMock()
        unauth = MagicMock(status_code=401)
        unauth.raise_for_status = MagicMock()
        success = MagicMock(status_code=200, json=lambda: {'status': 'PENDING', 'status_code': 200})
        success.raise_for_status = MagicMock()
        mock_post.side_effect = [token_resp, unauth, token_resp, success]

        client = BlastaClient('user', 'pass')
        status = client.check_dlr('12345')

    assert status == 'PENDING'


def test_check_dlr_raises_on_5xx():
    import requests as req_lib
    from modules.sms_processor import BlastaClient

    with patch('modules.sms_processor.requests.post') as mock_post:
        token_resp = MagicMock(status_code=200, json=lambda: {'access_token': 'tok'})
        token_resp.raise_for_status = MagicMock()
        server_err = MagicMock(status_code=500)
        server_err.raise_for_status = MagicMock(side_effect=req_lib.HTTPError('500'))
        mock_post.side_effect = [token_resp, server_err]

        client = BlastaClient('user', 'pass')
        with pytest.raises(req_lib.RequestException):
            client.check_dlr('12345')


# ---------------------------------------------------------------------------
# SmsProcessor.fetch_delivery_statuses
# ---------------------------------------------------------------------------

def _make_log_row(log_id, queue_id, subjid, provider_message_id):
    row = MagicMock()
    row._asdict.return_value = {
        'id': log_id,
        'queue_id': queue_id,
        'subjid': subjid,
        'provider_message_id': provider_message_id,
    }
    return row


def test_fetch_delivery_statuses_updates_terminal_status():
    from modules.sms_processor import SmsProcessor

    engine, conn = make_engine_mock(
        fetchall_return=[_make_log_row(1, 10, 'IBIS001', '274001')]
    )
    processor = SmsProcessor(config=make_config(), engine=engine)

    with patch.object(processor, '_get_client') as mock_get_client:
        mock_get_client.return_value.check_dlr.return_value = 'DELIVERED'
        result = processor.fetch_delivery_statuses()

    assert result.checked == 1
    assert result.updated == 1
    assert result.pending == 0
    assert result.errors == []


def test_fetch_delivery_statuses_leaves_pending_as_null():
    from modules.sms_processor import SmsProcessor

    engine, conn = make_engine_mock(
        fetchall_return=[_make_log_row(2, 20, 'IBIS002', '274002')]
    )
    processor = SmsProcessor(config=make_config(), engine=engine)

    with patch.object(processor, '_get_client') as mock_get_client:
        mock_get_client.return_value.check_dlr.return_value = 'PENDING'
        result = processor.fetch_delivery_statuses()

    assert result.checked == 1
    assert result.updated == 0
    assert result.pending == 1


def test_fetch_delivery_statuses_continues_on_single_error():
    from modules.sms_processor import SmsProcessor
    import requests as req_lib

    engine, conn = make_engine_mock(fetchall_return=[
        _make_log_row(3, 30, 'IBIS003', '274003'),
        _make_log_row(4, 40, 'IBIS004', '274004'),
    ])
    processor = SmsProcessor(config=make_config(), engine=engine)

    with patch.object(processor, '_get_client') as mock_get_client:
        mock_get_client.return_value.check_dlr.side_effect = [
            req_lib.RequestException('timeout'),
            'DELIVERED',
        ]
        result = processor.fetch_delivery_statuses()

    assert result.checked == 2
    assert result.updated == 1
    assert len(result.errors) == 1


def test_fetch_delivery_statuses_no_rows_returns_zero():
    from modules.sms_processor import SmsProcessor

    engine, conn = make_engine_mock(fetchall_return=[])
    processor = SmsProcessor(config=make_config(), engine=engine)
    result = processor.fetch_delivery_statuses()

    assert result.checked == 0
    assert result.updated == 0
